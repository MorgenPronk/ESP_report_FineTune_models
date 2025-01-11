##File name: excel_extraction.py

import pandas as pd
import pytesseract
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO
import logging_utils
import warnings
import logging
import re

# Suppress UserWarning from openpyxl mostly related to print areas
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Suppress lower-level logging messages from third-party libraries
logging.getLogger('openpyxl').setLevel(logging.ERROR)
logging.getLogger('pytesseract').setLevel(logging.ERROR)

# If Tesseract is not in your system PATH, you can explicitly provide the path
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Setup logging
# logger = logging_utils.setup_logger(f'logs/{__name__}.log')
logger = logging_utils.setup_logger(f'logs/main_log.log')

class ExcelExtractionError(Exception):
    """Custom exception for handling Excel extraction errors."""
    pass

# Function to perform OCR on images
def perform_ocr_on_image(image):
    return pytesseract.image_to_string(image)


# Function to extract text from Excel files
def get_text_from_excel(file_path):
    text = ""
    try:
        # Step 1: Extract text from cells
        if file_path.lower().endswith('.xls'):
            # Use xlrd engine for .xls files
            df = pd.read_excel(file_path, engine='xlrd')
        elif file_path.lower().endswith('.xlsx'):
            # Use openpyxl engine for .xlsx files
            df = pd.read_excel(file_path, engine='openpyxl')
        else:
            raise ExcelExtractionError(f"Unsupported Excel file format: {file_path} ")

        # Convert DataFrame to a string
        text += df.to_string()
    except Exception as e:
        logger.error(f"Error extracting text from Excel file {file_path}: {e}")
        raise ExcelExtractionError(f"Error extracting text from Excel file {file_path}: {e}")

    # Step 2: Extract images from .xlsx files
    if file_path.endswith('.xlsx'):
        try:
            text += extract_image_text_from_excel(file_path)
        except Exception as e:
            logger.error(f"Error extracting images from Excel file {file_path}: {e}")
            raise ExcelExtractionError(f"Error extracting images from Excel file {file_path}: {e}")

    # Step 3: remove "Unnamed" and "NaN" as a preprocess step
    processed_text = preprocess_text(text)

    return processed_text


# Function to extract images from Excel files and perform OCR
def extract_image_text_from_excel(file_path):
    text_from_images = ""
    try:
        # Load the Excel workbook
        wb = load_workbook(file_path, data_only=True)

        # Iterate through all worksheets
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]

            # Check if the worksheet contains any images
            if ws._images:
                for image in ws._images:
                    try:
                        # The image data is already in BytesIO, so no need to wrap it again.
                        img_data = image.ref  # or simply image.image for image data

                        # Open the image using PIL directly
                        img = Image.open(img_data)  # No need for BytesIO wrapping

                        # Perform OCR on the image
                        ocr_text = perform_ocr_on_image(img)
                        text_from_images += ocr_text

                    except Exception as e:
                        logger.error(f"Error processing image in Excel sheet {sheetname}: {e}")
                        raise ExcelExtractionError(f"Error processing image in Excel sheet {sheetname}: {e}")
    except Exception as e:
        logger.error(f"Error extracting images from Excel file {file_path}: {e}")
        raise ExcelExtractionError(f"Error extracting images from Excel file {file_path}: {e}")

    return text_from_images

def preprocess_text(text):
    """
    Cleans and reduces noise in the extracted text by handling placeholders like 'Unnamed' and 'NaN',
    and collapsing redundant whitespace.
    """

    # Replace "Unnamed" and similar placeholders
    text = re.sub(r'\bUnnamed\b.*?:?', '', text)  # Matches "Unnamed" or "Unnamed:X"

    # Replace "NaN" or similar placeholders
    text = re.sub(r'\bNaN\b', '', text)

    # Remove excessive whitespace introduced by replacements
    text = re.sub(r'\s+', ' ', text)  # Collapse multiple spaces into one
    text = re.sub(r'^\s+|\s+$', '', text, flags=re.MULTILINE)  # Strip leading/trailing spaces from each line

    # Remove rows that are completely empty after processing
    cleaned_lines = [line for line in text.splitlines() if line.strip()]
    text = "\n".join(cleaned_lines)

    return text
